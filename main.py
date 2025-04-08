import json
import pandas as pd
import numpy as np
from datetime import datetime
from fpdf import FPDF
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from scrapers.scraper1 import scrape_gym1
from scrapers.scraper2 import scrape_gym2
from scrapers.scraper3 import scrape_gym3
from scrapers.scraper4 import scrape_gym4
from scrapers.scraper5 import scrape_gym5
from scrapers.scraper6 import scrape_gym6
from scrapers.scraper7 import scrape_gym7
from scrapers.scraper8 import scrape_gym8
from scrapers.scraper9 import scrape_gym9
from scrapers.scraper10 import scrape_gym10
from scrapers.scraper11 import scrape_gym11
from scrapers.scraper12 import scrape_gym12
from scrapers.scraper13 import scrape_gym13
from scrapers.scraper14 import scrape_gym14
from scrapers.scraper15 import scrape_gym15

def main():
    print("🚀 Starting the Gym Scraper...\n")

    # **List of All Scrapers**
    scrapers = {
        "4x4 Squat Racks": [scrape_gym1, scrape_gym2, scrape_gym3, scrape_gym4, scrape_gym5],
        "Squat Stands": [scrape_gym6, scrape_gym7, scrape_gym8, scrape_gym9, scrape_gym10],
        "Leg Extensions": [scrape_gym11, scrape_gym12, scrape_gym13, scrape_gym14, scrape_gym15]
    }

    # **Initialize Lists for Data Storage**
    squat_racks = []
    squat_stands = []
    leg_extensions = []

    # **Run Scrapers and Collect Data**
    for category, scraper_list in scrapers.items():
        print(f"📌 Scraping {category}...\n")
        for scraper in scraper_list:
            try:
                data = scraper()

                # Ensure valid data is returned
                if not data or not isinstance(data, dict):
                    print(f"⚠️ {scraper.__name__} returned invalid data. Skipping.")
                    continue

                print(f"✅ Successfully Scraped: {data.get('name', 'Unknown')} - {data.get('price', 'N/A')}")

                # Append Data to the Correct List
                if category == "4x4 Squat Racks":
                    squat_racks.append(data)
                elif category == "Squat Stands":
                    squat_stands.append(data)
                elif category == "Leg Extensions":
                    leg_extensions.append(data)

            except Exception as e:
                print(f"❌ Error scraping {scraper.__name__}: {str(e)}")
                continue

    # **Save Data to Excel**
    print("\n📊 Saving Data to Excel...")
    save_to_excel_multi([
        (squat_racks, "4x4 Squat Racks"),
        (squat_stands, "Squat Stands"),
        (leg_extensions, "Leg Extensions")
    ], "gym_data.xlsx")

    # **Generate PDF Report**
    print("\n📄 Generating PDF Report...")
    generate_pdf()

    print("\n✅ All Tasks Completed Successfully!")

# -----------------------------------------------
# 🔧 Supporting Functions
# -----------------------------------------------

def format_excel(filename="gym_data.xlsx"):
    """Format the Excel file for better readability and fix hyperlinks."""
    wb = load_workbook(filename)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Set column widths for better readability
        column_widths = {"A": 25, "B": 20, "C": 15, "D": 12, "E": 25, "F": 35}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Bold headers and center align
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Fix hyperlinks
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[5].value and row[5].value.startswith("http"):
                row[5].hyperlink = row[5].value
                row[5].font = Font(color="0000FF", underline="single")

            if row[4].value and row[4].value.startswith("http"):
                row[4].hyperlink = row[4].value
                row[4].font = Font(color="0000FF", underline="single")

    # Save the formatted Excel file
    wb.save(filename)
    print(f"✅ Excel File Formatted & Links Fixed: {filename}")

def save_to_excel_multi(sheet_data_pairs, filename="gym_data.xlsx"):
    """Save scraped data to an Excel file with multiple sheets."""
    writer = pd.ExcelWriter(filename, engine="openpyxl")
    column_order = ["name", "manufacturer", "price", "country", "image_url", "web_page"]
    pretty_names = ["Product Name", "Manufacturer", "Price", "Country", "Image URL", "Product Page"]

    for data, sheet in sheet_data_pairs:
        df = pd.DataFrame(data)
        df = df[column_order] if not df.empty else pd.DataFrame(columns=column_order)
        df.fillna("NA", inplace=True)
        df.columns = pretty_names
        df.to_excel(writer, sheet_name=sheet, index=False)

    writer._save()
    print(f"✅ Data Saved to: {filename}")
    format_excel(filename)

def safe_text(text):
    """Ensure text is safe for PDF encoding by removing unsupported characters."""
    if pd.isna(text):  
        return "NA"
    return str(text).encode("latin-1", "ignore").decode("latin-1")

def generate_pdf(excel_file="gym_data.xlsx", pdf_file="gym_report.pdf"):
    """Generates a PDF report summarizing the gym equipment data."""
    xl = pd.ExcelFile(excel_file)
    report_date = datetime.now().strftime("%B %d, %Y")

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    # **Cover Page**
    pdf.add_page()
    pdf.set_font("Arial", style="B", size=24)
    pdf.cell(200, 20, safe_text("🏋️ Gym Equipment Report"), ln=True, align="C")
    pdf.ln(10)

    pdf.set_font("Arial", style="I", size=16)
    pdf.cell(200, 10, safe_text(f"📅 Last Updated: {report_date}"), ln=True, align="C")
    pdf.ln(20)

    pdf.set_font("Arial", size=14)
    pdf.multi_cell(0, 10, safe_text(
        "🔹 This report provides an up-to-date listing of gym equipment, "
        "including squat racks, stands, and leg extensions, along with pricing, manufacturer details, and product links.\n\n"
        "🔹 Data is collected from multiple leading gym equipment brands.\n\n"
        "🔹 Clickable product links are provided for quick access.\n\n"
        "📌 For the latest updates, please visit the respective manufacturer websites."
    ))

    pdf.ln(20)

    for sheet_name in xl.sheet_names:
        df = xl.parse(sheet_name)
        df.replace({np.nan: "NA"}, inplace=True)

        pdf.add_page()
        pdf.set_font("Arial", style="B", size=18)
        pdf.cell(200, 10, safe_text(sheet_name), ln=True, align="C")
        pdf.ln(10)

        for _, row in df.iterrows():
            pdf.set_font("Arial", style="B", size=12)
            pdf.cell(0, 10, safe_text(row["Product Name"]), ln=True)
            pdf.set_font("Arial", size=11)
            pdf.cell(0, 8, safe_text(f"💰 Price: {row['Price']}"), ln=True)
            pdf.cell(0, 8, safe_text(f"🏭 Manufacturer: {row['Manufacturer']}"), ln=True)
            pdf.cell(0, 8, safe_text(f"📍 Country: {row['Country']}"), ln=True)
            pdf.ln(10)

    pdf.output(pdf_file, "F")
    print(f"✅ PDF Report Generated: {pdf_file}")

if __name__ == "__main__":
    main()
